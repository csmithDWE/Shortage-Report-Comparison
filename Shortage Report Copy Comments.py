#Copies comments from one excel worksheet to another excel worksheet if the rows are otherwise identical
import tkinter as tk
from tkinter.filedialog import askdirectory
from tkinter import filedialog
from tkinter import messagebox
from tkinter import *
import os
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import *
from copy import copy

#File Handling and Setting Sheet 1 as active sheet for both workbooks
root = tk.Tk()
root.withdraw()
print("Select source file using file prompt:")
src_path = filedialog.askopenfilename()
print("Source File: \n"+src_path)
wb_obj_src = openpyxl.load_workbook(src_path)
sheet_obj_src = wb_obj_src["Sheet1"]
print("Select destination file using file prompt:")
dest_path = src_path = filedialog.askopenfilename()
print("Destination File: \n"+dest_path)
wb_obj_dest = openpyxl.load_workbook(dest_path)
sheet_obj_dest = wb_obj_dest["Sheet1"]
#

#Importing columns (import cols C,D,E from both sheets as vectors, match the entries, if an entry matches copy the comment to new sheet)
#To match the text boxes, compares every src to every destination to ensure that none are missed- ineffecient, do not use with large spreadsheets o(n2)
srcC = []
srcD = []
srcE = []
destC = []
destD = []
destE = []
srcT = []

for row in sheet_obj_src:
    val = row[2].value
    srcC.append(val)
for row in sheet_obj_src:
    val = row[3].value
    srcD.append(val)
for row in sheet_obj_src:
    val = row[4].value
    srcE.append(val)
for row in sheet_obj_dest:
    val = row[2].value
    destC.append(val)
for row in sheet_obj_dest:
    val = row[3].value
    destD.append(val)
for row in sheet_obj_dest:
    val = row[4].value
    destE.append(val)
# for row in sheet_obj_src:
#     val = row[19].value
#     srcT.append(val)
for i in range(1, sheet_obj_src.max_row+1):
    srcT.append(sheet_obj_src.cell(row = i, column = 20).value)



#create list that stores 1/0 for match or no match for a row's values in C/D/E, then use that vector at the end to compare the two comments cols
match = [None]*sheet_obj_src.max_row
for i in range(len(srcC)):
    for j in range(len(destC)):
        if(srcC[i] == destC[j] and destC[j] != None):
            if match[j] == None:
                match[j] = srcT[i]
                new_style = copy(sheet_obj_src.cell(row = i, column = 20).style)
                sheet_obj_dest.cell(row = j, column = 20).style = new_style

for i in range(len(srcD)):
    for j in range(len(destD)):
        if(srcD[i] == destD[j] and destD[j] != None):
            if match[j] == None:
                match[j] = srcT[i]
                new_style = copy(sheet_obj_src.cell(row = i, column = 20).style)
                sheet_obj_dest.cell(row = j, column = 20).style = new_style

for i in range(len(srcE)):
    for j in range(len(destE)):
        if(srcE[i] == destE[j] and destE[j] != None):
            if match[j] == None:
                match[j] = srcT[i]
                new_style = copy(sheet_obj_src.cell(row = i, column = 20).style)
                sheet_obj_dest.cell(row = j, column = 20).style = new_style


#Writing comments to dest sheet *ALWAYS WRITES TO COl19 (T)
for i in range(len(match)):
    if (match[i] != None):
        for j in range(1, sheet_obj_dest.max_row+1):
            sheet_obj_dest.cell(row = i+1, column = 20).value = match[i]
sheet_obj_dest.cell(row = 1, column = 20).value = "COPIED COMMENTS FROM PREVIOUS REPORT"






dest_path_new = dest_path.replace(".xlsx", "")
wb_obj_dest.save(dest_path_new + " w Comments.xlsx")
print("Operation Completed.")